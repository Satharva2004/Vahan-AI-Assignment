from __future__ import annotations

import argparse
import asyncio
import csv
import json
import mimetypes
import random
import sys
import wave
from dataclasses import asdict
from io import BytesIO
from pathlib import Path
from urllib.parse import urlparse

import httpx
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.metrics import evaluate
from app.providers import MODEL_SPECS, get_model, transcribe_audio


DEFAULT_MODELS = [model.id for model in MODEL_SPECS]


async def run(
    manifest_path: Path,
    output_dir: Path,
    models: list[str],
    frontend_public: Path | None = None,
    limit: int | None = None,
    sample_size: int | None = None,
    seed: int = 42,
) -> None:
    rows = load_rows(manifest_path)
    if sample_size:
        rows = random.Random(seed).sample(rows, min(sample_size, len(rows)))
    if limit:
        rows = rows[:limit]
    base_dir = manifest_path.parent
    output_dir.mkdir(parents=True, exist_ok=True)

    detailed = []
    summary_rows = []
    async with httpx.AsyncClient(timeout=180, follow_redirects=True) as client:
        for row_index, row in enumerate(rows, 1):
            content, filename, content_type = await load_audio(client, row["file"], base_dir)
            duration_seconds = audio_duration_seconds(content, content_type, filename)
            reference = row["reference"]
            entities = row.get("entities", "")

            for model_id in models:
                spec = get_model(model_id)
                print(f"[{row_index}/{len(rows)}] {filename} -> {spec.label}", flush=True)
                transcription = await transcribe_audio(spec, content, filename, content_type, {})
                metrics = asdict(evaluate(reference, transcription["transcript"], entities)) if transcription["ok"] else None
                rtf = real_time_factor(transcription["latency_ms"], duration_seconds) if transcription["ok"] else None
                record = {
                    "file": row["file"],
                    "file_name": filename,
                    "language": row.get("language", ""),
                    "condition": row.get("condition", "Unlabeled"),
                    "ground_truth": reference,
                    "entities": entities,
                    "duration_seconds": duration_seconds,
                    "model_id": spec.id,
                    "provider": spec.provider,
                    "model": spec.label,
                    "ok": transcription["ok"],
                    "latency_ms": transcription["latency_ms"],
                    "rtf": rtf,
                    "realtime_status": realtime_status(rtf),
                    "model_output": transcription["transcript"],
                    "error": transcription.get("error", ""),
                    "metrics": metrics,
                }
                detailed.append(record)
                summary_rows.append(flatten_record(record))

    write_outputs(output_dir, detailed, summary_rows, rows, models)
    if frontend_public:
        frontend_public.mkdir(parents=True, exist_ok=True)
        (frontend_public / "benchmark-results.json").write_text(
            json.dumps(build_frontend_payload(detailed, rows, models), indent=2, ensure_ascii=False),
            encoding="utf-8",
        )


def load_rows(manifest_path: Path) -> list[dict[str, str]]:
    if manifest_path.suffix.lower() in {".xlsx", ".xlsm"}:
        return load_excel_rows(manifest_path)
    with manifest_path.open(encoding="utf-8-sig", newline="") as handle:
        return [normalize_row(row) for row in csv.DictReader(handle)]


def load_excel_rows(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = [normalize_header(value) for value in rows[0]]
    parsed = []
    for row in rows[1:]:
        values = {headers[index]: stringify(value) for index, value in enumerate(row) if index < len(headers)}
        parsed.append(normalize_row(values))
    return [row for row in parsed if row.get("file") and row.get("reference")]


def normalize_row(row: dict) -> dict[str, str]:
    lowered = {normalize_header(key): stringify(value) for key, value in row.items()}
    return {
        "file": lowered.get("audio_file") or lowered.get("file") or lowered.get("audio") or "",
        "reference": lowered.get("ground_truth") or lowered.get("reference") or lowered.get("transcript") or "",
        "entities": lowered.get("entity") or lowered.get("entities") or "",
        "condition": lowered.get("condition") or "Unlabeled",
        "language": lowered.get("language") or lowered.get("langauge") or "",
    }


def normalize_header(value: object) -> str:
    return stringify(value).strip().lower().replace(" ", "_")


def stringify(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


async def load_audio(client: httpx.AsyncClient, file_value: str, base_dir: Path) -> tuple[bytes, str, str]:
    if file_value.startswith(("http://", "https://")):
        response = await client.get(file_value)
        response.raise_for_status()
        filename = Path(urlparse(file_value).path).name or "audio"
        content_type = response.headers.get("content-type", "").split(";", 1)[0]
        return response.content, filename, content_type or guess_content_type(filename)

    audio_path = (base_dir / file_value).resolve()
    return audio_path.read_bytes(), audio_path.name, guess_content_type(audio_path.name)


def guess_content_type(filename: str) -> str:
    return mimetypes.guess_type(filename)[0] or "application/octet-stream"


def audio_duration_seconds(content: bytes, content_type: str, filename: str) -> float | None:
    if content_type in {"audio/wav", "audio/x-wav", "audio/wave"} or filename.lower().endswith(".wav"):
        try:
            with wave.open(BytesIO(content), "rb") as audio:
                return round(audio.getnframes() / audio.getframerate(), 3)
        except (wave.Error, ZeroDivisionError):
            return None
    try:
        from mutagen import File as MutagenFile

        audio = MutagenFile(BytesIO(content))
        if audio is not None and audio.info and getattr(audio.info, "length", None):
            return round(float(audio.info.length), 3)
    except Exception:
        return None
    return None


def real_time_factor(latency_ms: int | float | None, duration_seconds: float | None) -> float | None:
    if not latency_ms or not duration_seconds or duration_seconds <= 0:
        return None
    return round((latency_ms / 1000) / duration_seconds, 3)


def realtime_status(rtf: float | None) -> str | None:
    if rtf is None:
        return None
    if rtf < 0.95:
        return "faster-than-real-time"
    if rtf <= 1.05:
        return "real-time"
    return "slower-than-real-time"


def flatten_record(record: dict) -> dict:
    metrics = record["metrics"] or {}
    return {
        "file_name": record["file_name"],
        "language": record["language"],
        "condition": record["condition"],
        "ground_truth": record["ground_truth"],
        "entities": record["entities"],
        "model": record["model"],
        "ok": record["ok"],
        "wer": metrics.get("wer", ""),
        "accuracy": max(0, 1 - metrics["wer"]) if metrics.get("wer") is not None else "",
        "cer": metrics.get("cer", ""),
        "exact_match": metrics.get("wer") == 0 if metrics.get("wer") is not None else "",
        "entity_recall": metrics.get("entity_recall", ""),
        "entity_f1": metrics.get("entity_f1", ""),
        "hallucination_rate": metrics.get("hallucination_rate", ""),
        "latency_ms": record["latency_ms"],
        "rtf": record["rtf"] or "",
        "realtime_status": record["realtime_status"] or "",
        "missed_entities": "; ".join(metrics.get("missed_entities", [])),
        "model_output": record["model_output"],
        "error": record["error"],
    }


def write_outputs(output_dir: Path, detailed: list[dict], summary_rows: list[dict], source_rows: list[dict], models: list[str]) -> None:
    (output_dir / "results.json").write_text(json.dumps(detailed, indent=2, ensure_ascii=False), encoding="utf-8")
    with (output_dir / "results.csv").open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(summary_rows[0].keys()))
        writer.writeheader()
        writer.writerows(summary_rows)

    aggregates = build_aggregates(detailed, source_rows, models)
    (output_dir / "summary.json").write_text(json.dumps(aggregates, indent=2, ensure_ascii=False), encoding="utf-8")


def build_aggregates(detailed: list[dict], source_rows: list[dict], models: list[str]) -> list[dict]:
    aggregates = []
    for model_id in models:
        spec = get_model(model_id)
        ok_records = [item for item in detailed if item["model_id"] == model_id and item["ok"]]
        metric_records = [item for item in ok_records if item["metrics"]]
        aggregates.append(
            {
                "model_id": spec.id,
                "provider": spec.provider,
                "model": spec.label,
                "recordings": len(ok_records),
                "wer": avg(metric_records, "wer"),
                "accuracy": accuracy_from_wer(avg(metric_records, "wer")),
                "cer": avg(metric_records, "cer"),
                "exact_match_rate": exact_match_rate(metric_records),
                "entity_recall": avg(metric_records, "entity_recall"),
                "entity_f1": avg(metric_records, "entity_f1"),
                "hallucination_rate": avg(metric_records, "hallucination_rate"),
                "latency_ms": round(sum(item["latency_ms"] for item in ok_records) / len(ok_records)) if ok_records else None,
                "rtf": avg_raw(ok_records, "rtf"),
                "failures": len(source_rows) - len(ok_records),
            }
        )
    return sorted(aggregates, key=lambda item: (item["wer"] is None, item["wer"] or 999))


def build_frontend_payload(detailed: list[dict], source_rows: list[dict], models: list[str]) -> dict:
    return {
        "source": "data/Voice Notes.xlsx",
        "recording_count": len(source_rows),
        "models": build_aggregates(detailed, source_rows, models),
        "rows": detailed,
    }


def avg(records: list[dict], key: str) -> float | None:
    if not records:
        return None
    return round(sum(item["metrics"][key] for item in records) / len(records), 4)


def accuracy_from_wer(wer: float | None) -> float | None:
    if wer is None:
        return None
    return round(max(0, 1 - wer), 4)


def exact_match_rate(records: list[dict]) -> float | None:
    if not records:
        return None
    return round(sum(1 for item in records if item["metrics"]["wer"] == 0) / len(records), 4)


def avg_raw(records: list[dict], key: str) -> float | None:
    values = [item[key] for item in records if item.get(key) is not None]
    if not values:
        return None
    return round(sum(values) / len(values), 3)


def main() -> None:
    parser = argparse.ArgumentParser(description="Run ASR benchmark from CSV or Excel.")
    parser.add_argument("--manifest", default="../data/manifest.csv", help="Path to manifest CSV/XLSX.")
    parser.add_argument("--output", default="../data/results", help="Output directory.")
    parser.add_argument("--frontend-public", default="../frontend/public", help="Optional frontend public directory.")
    parser.add_argument("--limit", type=int, default=None, help="Optional row limit for smoke tests.")
    parser.add_argument("--sample-size", type=int, default=None, help="Random sample size.")
    parser.add_argument("--seed", type=int, default=42, help="Random sample seed.")
    parser.add_argument("--models", nargs="*", default=DEFAULT_MODELS, help="Model ids to run.")
    args = parser.parse_args()
    asyncio.run(
        run(
            Path(args.manifest).resolve(),
            Path(args.output).resolve(),
            args.models,
            Path(args.frontend_public).resolve() if args.frontend_public else None,
            args.limit,
            args.sample_size,
            args.seed,
        )
    )


if __name__ == "__main__":
    main()
